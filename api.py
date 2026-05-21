"""
api.py — Mapa de Contagem de Tráfego
Fotos servidas via Google Drive API (pastas públicas)
"""

import os, json, urllib.request, urllib.parse, zipfile, io, tempfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from datetime import datetime, timezone, timedelta
BRASILIA = timezone(timedelta(hours=-3))
from typing import Optional
from functools import lru_cache
import time

from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.responses import JSONResponse, HTMLResponse
from fastapi.middleware.cors import CORSMiddleware

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    raise SystemExit("Instale: pip install pandas openpyxl")

# ── Configuração ──────────────────────────────────────────────────────────────
BASE_DIR  = Path(__file__).parent
HTML_PATH       = BASE_DIR / "mapa_online.html"
FINANCEIRO_PATH = BASE_DIR / "financeiro.xlsx"

def encontrar_excel():
    candidatos = [
        "PONTOS_DE_CONTAGEM_DE_TR\u00c1FEGO.xlsx",
        "PONTOS DE CONTAGEM DE TR\u00c1FEGO.xlsx",
        "PONTOS_DE_CONTAGEM_DE_TRAFEGO.xlsx",
        "PONTOS DE CONTAGEM DE TRAFEGO.xlsx",
    ]
    for nome in candidatos:
        p = BASE_DIR / nome
        if p.exists():
            return p
    xlsx = list(BASE_DIR.glob("*.xlsx"))
    if xlsx:
        return xlsx[0]
    return BASE_DIR / "PONTOS_DE_CONTAGEM_DE_TR\u00c1FEGO.xlsx"

EXCEL_PATH = encontrar_excel()

ABA_PONTOS      = "LISTA DE PONTOS"
ABA_DADOS       = "DADOS COLETADOS"

# Google Drive
DRIVE_API_KEY       = os.environ.get("DRIVE_API_KEY", "AIzaSyATY3YHzzMJTB92heZcIxIJBGpK5Y2Ff6A")
DRIVE_ROOT_FOLDER   = os.environ.get("DRIVE_ROOT_FOLDER", "1yrBYApnWeKEe7YEJT7I6Clc5JgeDFpCj")
DRIVE_BASE_URL      = "https://www.googleapis.com/drive/v3"
DRIVE_VIEW_URL      = "https://lh3.googleusercontent.com/d/"

# Cache de 10 minutos para não bater na API a cada clique
_cache_fotos: dict = {}
_cache_subpastas: dict = {}
_cache_timestamp: dict = {}
CACHE_TTL = 600  # segundos

CLASSES = ['MOTO','PASS','UTIL','2C/2CB','3C/3CB',
           '2S1','2S2','2S3','3S1','3S2','3S3',
           '4C','4CD','2C2','2C3','3C2','3C3','2I3','3I3','BIT','ROD/TRIT']

EXTS_FOTO = {'image/jpeg','image/png','image/webp','image/jpg'}

# ── App ───────────────────────────────────────────────────────────────────────
app = FastAPI(title="Mapa de Contagem de Tráfego", version="2.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

# ── Google Drive helpers ──────────────────────────────────────────────────────
def drive_list(folder_id: str) -> list:
    """Lista arquivos/pastas dentro de uma pasta do Drive, paginando todos os resultados."""
    q = urllib.parse.quote(f"'{folder_id}' in parents and trashed=false")
    fields = urllib.parse.quote("nextPageToken,files(id,name,mimeType)")
    all_files = []
    page_token = ""
    while True:
        token_param = f"&pageToken={urllib.parse.quote(page_token)}" if page_token else ""
        url = f"{DRIVE_BASE_URL}/files?q={q}&key={DRIVE_API_KEY}&fields={fields}&pageSize=1000{token_param}"
        try:
            resp = urllib.request.urlopen(url, timeout=15)
            data = json.loads(resp.read())
            all_files.extend(data.get("files", []))
            page_token = data.get("nextPageToken", "")
            if not page_token:
                break
        except Exception as e:
            print(f"Drive API erro: {e}")
            break
    return all_files

def encontrar_pasta_id(parent_id: str, nome_busca: str) -> Optional[str]:
    """Encontra ID de uma subpasta pelo nome (case-insensitive, aceita variações)."""
    items = drive_list(parent_id)
    # Tenta match exato primeiro
    for item in items:
        if item["mimeType"] == "application/vnd.google-apps.folder":
            if item["name"].upper() == nome_busca.upper():
                return item["id"]
    # Tenta match parcial com "instal"
    for item in items:
        if item["mimeType"] == "application/vnd.google-apps.folder":
            if "instal" in item["name"].lower():
                return item["id"]
    # Se só tiver uma subpasta, usa ela
    pastas = [i for i in items if i["mimeType"] == "application/vnd.google-apps.folder"]
    if len(pastas) == 1:
        return pastas[0]["id"]
    return None

def listar_fotos_drive(pid: str) -> list:
    """Retorna lista de URLs diretas das fotos do ponto no Drive."""
    now = time.time()

    # Cache válido?
    if pid in _cache_fotos and (now - _cache_timestamp.get(pid, 0)) < CACHE_TTL:
        return _cache_fotos[pid]

    # 1. Encontra pasta do ponto (ex: "E 05") dentro da raiz
    if "root_items" not in _cache_subpastas or (now - _cache_timestamp.get("root", 0)) > CACHE_TTL:
        items = drive_list(DRIVE_ROOT_FOLDER)
        _cache_subpastas["root_items"] = {
            i["name"]: i["id"] for i in items
            if i["mimeType"] == "application/vnd.google-apps.folder"
        }
        _cache_timestamp["root"] = now

    pasta_ponto_id = _cache_subpastas["root_items"].get(pid)
    if not pasta_ponto_id:
        _cache_fotos[pid] = []
        _cache_timestamp[pid] = now
        return []

    # 2. Encontra subpasta FOTOS - INSTALAÇÃO
    sub_id = encontrar_pasta_id(pasta_ponto_id, "FOTOS - INSTALAÇÃO")
    if not sub_id:
        sub_id = pasta_ponto_id  # fotos direto na pasta do ponto

    # 3. Lista arquivos de imagem
    items = drive_list(sub_id)
    fotos = sorted(
        [i for i in items if i.get("mimeType","") in EXTS_FOTO],
        key=lambda x: x["name"].lower()
    )
    urls = [DRIVE_VIEW_URL + f["id"] for f in fotos]

    _cache_fotos[pid] = urls
    _cache_timestamp[pid] = now
    return urls

# ── Helpers Excel ─────────────────────────────────────────────────────────────
def limpar(v):
    return "" if (v is None or str(v).strip() == "nan") else str(v).strip()

def coord(v):
    try: return float(str(v).replace(",", ".").strip())
    except: return None

def fmt_data(v):
    if pd.isna(v): return None
    try:
        ts = pd.Timestamp(v)
        return None if ts.year < 1950 else ts.strftime("%d/%m/%Y")
    except: return None

def ler_excel():
    if not EXCEL_PATH.exists():
        raise HTTPException(404, f"Excel não encontrado: {EXCEL_PATH.name}")

    df = pd.read_excel(EXCEL_PATH, sheet_name=ABA_PONTOS, dtype=str)
    df.dropna(how="all", inplace=True)

    col_map = {c.upper().strip(): c for c in df.columns}
    def col(cands):
        for c in cands:
            if c.upper() in col_map: return col_map[c.upper()]
        return None

    c_id  = col(["ID"])
    c_lat = col(["LATITUDE","LAT"])
    c_lng = col(["LONGITUDE","LNG","LONG"])
    c_rod = col(["RODOVIA"])
    c_tre = col(["TRECHO"])
    c_ini = col(["DESCRIÇÃO INÍCIO","DESCRICAO INICIO","INICIO"])
    c_fim = col(["DESCRIÇÃO FINAL","DESCRICAO FINAL","FIM"])
    c_sug = col(["SUGESTÃO","SUGESTAO","RESPONSAVEL","RESPONSÁVEL"])
    c_sta = col(["STATUS DE EXECUÇÃO","STATUS DE EXECUCAO","STATUS"])
    c_pis = col(["SITUAÇÃO PISTA","SITUACAO PISTA","PISTA"])

    # Dados coletados
    dados_dc = {}
    try:
        dc = pd.read_excel(EXCEL_PATH, sheet_name=ABA_DADOS, header=None)
        for i in range(4, len(dc)):
            row = dc.iloc[i]
            pid = limpar(str(row[0])) if pd.notna(row[0]) else ""
            if not pid or pid == "nan": continue
            total = int(row[28]) if pd.notna(row[28]) and str(row[28]) not in ("nan","") else 0
            contagens = {}
            for j, cls in enumerate(CLASSES):
                val = row[7+j]
                contagens[cls] = int(val) if pd.notna(val) and str(val) not in ("nan","") else 0
            dados_dc[pid] = {
                "periodo_inicio": fmt_data(row[5]),
                "periodo_fim":    fmt_data(row[6]),
                "total": total,
                "contagens": contagens,
            }
    except Exception as e:
        print(f"Aviso dados coletados: {e}")

    pontos = []
    for _, row in df.iterrows():
        pid = limpar(row.get(c_id, "")) if c_id else ""
        if not pid: continue
        lat = coord(row.get(c_lat, "")) if c_lat else None
        lng = coord(row.get(c_lng, "")) if c_lng else None
        if lat is None or lng is None: continue

        def g(c): return limpar(row.get(c, "")) if c else ""
        dc = dados_dc.get(pid, {})

        pontos.append({
            "id":       pid,
            "lat":      lat,
            "lng":      lng,
            "rodovia":  g(c_rod),
            "trecho":   g(c_tre),
            "inicio":   g(c_ini),
            "fim":      g(c_fim),
            "sugestao": g(c_sug) or "-",
            "status":   g(c_sta) or "-",
            "pista":    g(c_pis) or "-",
            "fotos":    [],  # carregado sob demanda via /api/fotos/{pid}
            "total":           dc.get("total", 0),
            "periodo_inicio":  dc.get("periodo_inicio"),
            "periodo_fim":     dc.get("periodo_fim"),
            "contagens":       dc.get("contagens", {}),
        })

    return pontos

# ── Rotas ─────────────────────────────────────────────────────────────────────
@app.get("/", response_class=HTMLResponse)
def raiz():
    if HTML_PATH.exists():
        return HTMLResponse(HTML_PATH.read_text(encoding="utf-8"))
    return HTMLResponse("<h2>mapa.html não encontrado</h2>")

# ── Edição de status via GitHub API (token fica seguro no servidor) ────────────
GH_TOKEN_ENV  = os.environ.get("GH_TOKEN", "")
GH_REPO_ENV   = "lleollopes-web/mapa-contagem"
GH_FILE_ENV   = "status.json"
GH_BRANCH_ENV = "principal"
GH_API_BASE   = "https://api.github.com"

STATUS_VALIDOS = {"CONCLUIDO", "EM ANDAMENTO", "NAO INICIADO"}

from pydantic import BaseModel

class StatusUpdate(BaseModel):
    id: str
    status: str

@app.post("/api/status/update")
def update_status(body: StatusUpdate):
    if not GH_TOKEN_ENV:
        raise HTTPException(500, "GH_TOKEN não configurado no servidor")
    if body.status.upper() not in STATUS_VALIDOS:
        raise HTTPException(400, f"Status inválido: {body.status}")

    headers = {
        "Authorization": f"token {GH_TOKEN_ENV}",
        "Content-Type": "application/json",
        "Accept": "application/vnd.github+json",
    }

    # 1. Busca conteúdo atual do status.json
    url_file = f"{GH_API_BASE}/repos/{GH_REPO_ENV}/contents/{GH_FILE_ENV}"
    try:
        req = urllib.request.Request(url_file, headers=headers)
        resp = urllib.request.urlopen(req, timeout=15)
        file_data = json.loads(resp.read())
        sha = file_data["sha"]
        import base64 as b64mod
        content_raw = file_data["content"].replace("\n","").replace(" ","")
        try:
            conteudo_atual = json.loads(b64mod.b64decode(content_raw).decode("utf-8"))
        except:
            conteudo_atual = {}
        print(f"  status.json lido, sha={sha}, entries={len(conteudo_atual)}")
    except urllib.error.HTTPError as e:
        body_err = e.read().decode("utf-8", errors="replace")
        print(f"  Erro HTTP ao ler status.json: {e.code} {body_err}")
        if e.code == 404:
            sha = None
            conteudo_atual = {}
        else:
            raise HTTPException(502, f"Erro ao ler status.json: {e.code} {body_err}")
    except Exception as e:
        print(f"  Erro ao ler status.json: {e}")
        raise HTTPException(502, f"Erro ao ler status.json: {e}")

    # 2. Atualiza e faz commit
    conteudo_atual[body.id] = body.status.upper()
    novo_conteudo = json.dumps(conteudo_atual, ensure_ascii=False, indent=2)
    import base64 as b64mod2
    conteudo_b64 = b64mod2.b64encode(novo_conteudo.encode("utf-8")).decode("ascii")

    payload = {
        "message": f"Status: {body.id} → {body.status.upper()}",
        "content": conteudo_b64,
    }
    if sha:
        payload["sha"] = sha

    try:
        req2 = urllib.request.Request(
            f"{GH_API_BASE}/repos/{GH_REPO_ENV}/contents/{GH_FILE_ENV}",
            data=json.dumps(payload).encode("utf-8"),
            headers=headers,
            method="PUT",
        )
        urllib.request.urlopen(req2, timeout=15)
    except urllib.error.HTTPError as e:
        detail = e.read().decode("utf-8", errors="replace")
        raise HTTPException(502, f"Erro ao salvar: {e.code} {detail}")
    except Exception as e:
        raise HTTPException(502, f"Erro ao salvar: {e}")

    return JSONResponse({"ok": True, "id": body.id, "status": body.status.upper()})

@app.get("/api/pontos")
def get_pontos():
    pontos = ler_excel()
    excel = encontrar_excel()
    try:
        excel_mtime = datetime.fromtimestamp(excel.stat().st_mtime, tz=BRASILIA).strftime("%d/%m/%Y %H:%M")
    except:
        excel_mtime = datetime.now(BRASILIA).strftime("%d/%m/%Y %H:%M")
    return JSONResponse({
        "total": len(pontos),
        "com_contagem": sum(1 for p in pontos if p["total"] > 0),
        "atualizado_em": excel_mtime,
        "pontos": pontos,
    })

@app.get("/api/ponto/{pid}")
def get_ponto(pid: str):
    pontos = ler_excel()
    for p in pontos:
        if p["id"] == pid: return JSONResponse(p)
    raise HTTPException(404, f"Ponto '{pid}' não encontrado")

def encontrar_xlsx_drive(pid: str) -> Optional[str]:
    """Encontra o arquivo .xlsx na pasta do ponto no Drive e retorna URL de download."""
    now = time.time()
    # Reutiliza cache de subpastas
    if "root_items" not in _cache_subpastas or (now - _cache_timestamp.get("root", 0)) > CACHE_TTL:
        items = drive_list(DRIVE_ROOT_FOLDER)
        _cache_subpastas["root_items"] = {
            i["name"]: i["id"] for i in items
            if i["mimeType"] == "application/vnd.google-apps.folder"
        }
        _cache_timestamp["root"] = now
    pasta_ponto_id = _cache_subpastas["root_items"].get(pid)
    if not pasta_ponto_id:
        return None
    # Lista arquivos direto na pasta do ponto (não subpasta)
    items = drive_list(pasta_ponto_id)
    for item in items:
        if (item.get("mimeType","") in
            {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
             "application/vnd.ms-excel"}
            or item.get("name","").endswith(".xlsx")):
            return f"https://drive.google.com/uc?export=download&id={item['id']}"
    return None

@app.get("/api/fotos/{pid}")
def get_fotos(pid: str):
    """Retorna fotos e link do xlsx de um ponto sob demanda"""
    fotos = listar_fotos_drive(pid)
    xlsx  = encontrar_xlsx_drive(pid)
    return JSONResponse({"pid": pid, "fotos": fotos, "total": len(fotos), "xlsx": xlsx})

@app.post("/api/upload-excel")
async def upload_excel(file: UploadFile = File(...)):
    if not file.filename.endswith((".xlsx",".xls")):
        raise HTTPException(400, "Apenas .xlsx aceito")
    EXCEL_PATH.write_bytes(await file.read())
    # Limpa cache de fotos
    _cache_fotos.clear()
    _cache_subpastas.clear()
    _cache_timestamp.clear()
    return {"ok": True, "atualizado_em": datetime.now(BRASILIA).strftime("%d/%m/%Y %H:%M:%S")}

@app.get("/api/limpar-cache")
def limpar_cache():
    """Força releitura das fotos do Drive"""
    _cache_fotos.clear()
    _cache_subpastas.clear()
    _cache_timestamp.clear()
    return {"ok": True, "msg": "Cache limpo"}

@app.get("/api/download/zip-xlsx")
def download_zip_xlsx():
    """Baixa ZIP com todos os xlsx dos pontos concluidos — download paralelo"""
    from fastapi.responses import StreamingResponse
    import unicodedata

    pontos = ler_excel()
    concluidos = []
    for p in pontos:
        k = unicodedata.normalize("NFD", p.get("status","").upper())
        k = "".join(c for c in k if unicodedata.category(c) != "Mn")
        if "CONCLU" in k:
            concluidos.append(p)

    if not concluidos:
        raise HTTPException(404, "Nenhum ponto concluido encontrado")

    # Garante cache do root
    now = time.time()
    if "root_items" not in _cache_subpastas or (now - _cache_timestamp.get("root", 0)) > CACHE_TTL:
        items = drive_list(DRIVE_ROOT_FOLDER)
        _cache_subpastas["root_items"] = {
            i["name"]: i["id"] for i in items
            if i["mimeType"] == "application/vnd.google-apps.folder"
        }
        _cache_timestamp["root"] = now

    def baixar_ponto(p):
        """Retorna (nome_arquivo, bytes) ou None"""
        pid = p["id"]
        pasta_id = _cache_subpastas["root_items"].get(pid)
        if not pasta_id:
            return None
        arquivos = drive_list(pasta_id)
        for arq in arquivos:
            mime = arq.get("mimeType", "")
            nome = arq.get("name", "")
            if (mime in {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         "application/vnd.ms-excel"}
                    or nome.lower().endswith(".xlsx")):
                try:
                    url = f"{DRIVE_BASE_URL}/files/{arq['id']}?alt=media&key={DRIVE_API_KEY}"
                    req = urllib.request.urlopen(url, timeout=30)
                    dados = req.read()
                    return (f"{pid}/{nome}", dados)
                except Exception as e:
                    print(f"Erro {pid}: {e}")
        return None

    # Download paralelo — até 8 pontos simultaneamente
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        with ThreadPoolExecutor(max_workers=8) as executor:
            futures = {executor.submit(baixar_ponto, p): p for p in concluidos}
            for future in as_completed(futures):
                result = future.result()
                if result:
                    nome_arq, dados = result
                    zf.writestr(nome_arq, dados)

    zip_buffer.seek(0)
    from datetime import datetime as dt
    nome_zip = f"Contagens_Concluidas_{dt.now(BRASILIA).strftime('%d%m%Y')}.zip"
    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename={nome_zip}"}
    )


@app.get("/api/download/resumo-xlsx")
def download_resumo():
    """Gera planilha resumo dos pontos concluidos"""
    from fastapi.responses import StreamingResponse
    from openpyxl.utils import get_column_letter
    from datetime import datetime as dt
    import unicodedata

    pontos = ler_excel()
    concluidos = []
    for p in pontos:
        k = unicodedata.normalize("NFD", p.get("status","").upper())
        k = "".join(c for c in k if unicodedata.category(c) != "Mn")
        if "CONCLU" in k:
            concluidos.append(p)

    if not concluidos:
        raise HTTPException(404, "Nenhum ponto concluido encontrado")

    wb = Workbook()
    ws = wb.active
    ws.title = "Pontos Concluidos"

    header_fill = PatternFill("solid", fgColor="1565C0")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left        = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin        = Side(style="thin", color="CCCCCC")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    green_fill  = PatternFill("solid", fgColor="E8F5E9")
    green_font  = Font(bold=True, color="1B5E20", size=11)

    NCOLS = 10

    ws.merge_cells(f"A1:{get_column_letter(NCOLS)}1")
    ws["A1"] = "RELATORIO DE PONTOS DE CONTAGEM CONCLUIDOS"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=13)
    ws["A1"].fill = PatternFill("solid", fgColor="0D47A1")
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f"A2:{get_column_letter(NCOLS)}2")
    ws["A2"] = f"COMOL Consultoria M.L.  |  SOP-CE  |  Gerado em: {dt.now(BRASILIA).strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(italic=True, color="555555", size=10)
    ws["A2"].fill = PatternFill("solid", fgColor="E3F2FD")
    ws["A2"].alignment = center
    ws.row_dimensions[2].height = 18

    headers = ["N", "ID Ponto", "Codigo Trecho", "Rodovia",
               "Descricao Inicio", "Descricao Fim",
               "Latitude", "Longitude",
               "Periodo Inicio", "Periodo Fim"]
    ws.row_dimensions[3].height = 36
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    widths = [5, 10, 16, 10, 36, 36, 13, 13, 13, 13]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    alt_fill = PatternFill("solid", fgColor="F5F5F5")
    for idx, p in enumerate(concluidos, 1):
        row = idx + 3
        ws.row_dimensions[row].height = 22
        fill = alt_fill if idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        valores = [
            idx,
            p.get("id", ""),
            p.get("trecho", ""),
            p.get("rodovia", ""),
            p.get("inicio", ""),
            p.get("fim", ""),
            p.get("lat", ""),
            p.get("lng", ""),
            p.get("periodo_inicio") or "-",
            p.get("periodo_fim")    or "-",
        ]
        for col, val in enumerate(valores, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill
            cell.border = border
            cell.alignment = center if col in [1,2,3,4,7,8,9,10] else left
            cell.font = Font(size=10)

    tot_row = len(concluidos) + 4
    ws.merge_cells(f"A{tot_row}:{get_column_letter(NCOLS)}{tot_row}")
    ws[f"A{tot_row}"] = f"TOTAL DE PONTOS CONCLUIDOS: {len(concluidos)}"
    ws[f"A{tot_row}"].font = green_font
    ws[f"A{tot_row}"].fill = green_fill
    ws[f"A{tot_row}"].alignment = center

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    nome = f"Resumo_Contagens_{dt.now(BRASILIA).strftime('%d%m%Y')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nome}"}
    )


@app.get("/api/download/vmd-xlsx")
def download_vmd():
    """Gera planilha VDM com todas as classes de veiculos dos pontos com dados coletados"""
    from fastapi.responses import StreamingResponse
    from openpyxl.utils import get_column_letter
    from datetime import datetime as dt

    df_raw   = pd.read_excel(EXCEL_PATH, sheet_name=ABA_DADOS,  header=None)
    df_lista = pd.read_excel(EXCEL_PATH, sheet_name=ABA_PONTOS)

    CLASSES = ["MOTO","PASS","UTIL","2C/2CB","3C/3CB",
               "2S1","2S2","2S3","3S1","3S2","3S3",
               "4C","4CD","2C2","2C3","3C2","3C3","2I3","3I3","BIT","ROD/TRIT"]

    mask = pd.to_numeric(df_raw[28], errors="coerce") > 0
    dados = df_raw[mask].copy()
    if dados.empty:
        raise HTTPException(404, "Nenhum ponto com dados coletados encontrado")

    rows = []
    for _, row in dados.iterrows():
        pid         = str(row[0]).strip()
        info        = df_lista[df_lista["ID"].astype(str).str.strip() == pid]
        rodovia     = str(row[1]) if pd.notna(row[1]) else ""
        trecho      = str(row[2]) if pd.notna(row[2]) else ""
        inicio_desc = info["DESCRICAO INICIO"].values[0] if len(info) and "DESCRICAO INICIO" in info.columns else (info["DESCRI\u00c7\u00c3O IN\u00cdCIO"].values[0] if len(info) else "")
        fim_desc    = info["DESCRICAO FINAL"].values[0]  if len(info) and "DESCRICAO FINAL" in info.columns else (info["DESCRI\u00c7\u00c3O FINAL"].values[0] if len(info) else "")
        periodo_ini = pd.to_datetime(row[5]).strftime('%d/%m/%Y') if pd.notna(row[5]) else "-"
        periodo_fim = pd.to_datetime(row[6]).strftime('%d/%m/%Y') if pd.notna(row[6]) else "-"
        vals        = [int(row[7+i]) if pd.notna(row[7+i]) and str(row[7+i]) not in ["nan",""] else 0
                       for i in range(21)]
        total       = int(row[28]) if pd.notna(row[28]) else 0
        leves    = vals[0]+vals[1]+vals[2]
        onibus   = vals[3]+vals[4]
        caminhao = sum(vals[5:])
        periodo = f"{periodo_ini} – {periodo_fim}" if periodo_ini != "-" else "-"
        rows.append([pid, rodovia, trecho, inicio_desc, fim_desc,
                     periodo] + vals + [leves, onibus, caminhao, total])

    def hfill(cor): return PatternFill("solid", fgColor=cor)
    thin = Side(style="thin", color="CCCCCC")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lft  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    NCOLS = 32

    wb = Workbook(); ws = wb.active; ws.title = "VDM Trechos"

    ws.merge_cells(f"A1:{get_column_letter(NCOLS)}1")
    ws["A1"] = "PLANILHA RESUMO - VDM TRECHOS CONCLUIDOS"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=13)
    ws["A1"].fill = hfill("0D47A1"); ws["A1"].alignment = ctr
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f"A2:{get_column_letter(NCOLS)}2")
    ws["A2"] = f"COMOL Consultoria M.L.  |  SOP-CE  |  Gerado em: {dt.now(BRASILIA).strftime('%d/%m/%Y %H:%M')}  |  {len(rows)} pontos"
    ws["A2"].font = Font(italic=True, color="555555", size=10)
    ws["A2"].fill = hfill("E3F2FD"); ws["A2"].alignment = ctr
    ws.row_dimensions[2].height = 18

    for ini, fim, txt, cor in [
        ("A3","G3","","1565C0"), ("H3","J3","VEICULOS LEVES","1976D2"),
        ("K3","L3","ONIBUS","7B1FA2"), ("M3","W3","CAMINHAO","E65100"),
        ("X3","Z3","SUBTOTAIS","2E7D32"), ("AA3","AF3","VDM TOTAL","0D47A1")]:
        ws.merge_cells(f"{ini}:{fim}")
        c = ws[ini]; c.value = txt
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = hfill(cor); c.alignment = ctr; c.border = brd
    ws.row_dimensions[3].height = 20

    for ci, h in enumerate(["N","ID","Rodovia","Trecho","Descricao Inicio","Descricao Fim",
            "Periodo"] + CLASSES + ["Leves","Onibus","Caminhao","VDM TOTAL"], 1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = hfill("1565C0"); c.alignment = ctr; c.border = brd
    ws.row_dimensions[4].height = 36

    for i, w in enumerate([5,8,10,16,34,34,22]+[7]*21+[9,9,11,11], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    totais_cls = [0]*21
    total_leves = total_onibus = total_cam = total_geral = 0
    for idx, r in enumerate(rows, 1):
        rn = idx + 4; ws.row_dimensions[rn].height = 18
        bg = hfill("F5F5F5") if idx % 2 == 0 else hfill("FFFFFF")
        vals = r[6:27]; leves = r[27]; onibus = r[28]; cam = r[29]; total = r[30]
        for ci, val in enumerate([idx]+r[:6]+vals+[leves,onibus,cam,total], 1):
            c = ws.cell(row=rn, column=ci, value=val)
            c.border = brd; c.font = Font(size=10)
            c.alignment = lft if ci in [5,6] else ctr
            if ci == 32: c.font = Font(bold=True, color="0D47A1", size=10); c.fill = hfill("E3F2FD")
            elif ci in [29,30,31]: c.font = Font(bold=True, color="1B5E20", size=10); c.fill = hfill("E8F5E9")
            else: c.fill = bg
        for i, v in enumerate(vals): totais_cls[i] += v
        total_leves += leves; total_onibus += onibus; total_cam += cam; total_geral += total

    tr = len(rows) + 5; ws.row_dimensions[tr].height = 22
    ws.merge_cells(f"A{tr}:G{tr}")
    ws[f"A{tr}"] = f"TOTAL GERAL - {len(rows)} pontos"
    ws[f"A{tr}"].font = Font(bold=True, color="1B5E20", size=11)
    ws[f"A{tr}"].fill = hfill("E8F5E9"); ws[f"A{tr}"].alignment = ctr; ws[f"A{tr}"].border = brd
    for i, v in enumerate(totais_cls, 8):
        c = ws.cell(row=tr, column=i, value=v)
        c.font = Font(bold=True, color="333333", size=10)
        c.fill = hfill("E8F5E9"); c.alignment = ctr; c.border = brd
    for ci, val in [(29,total_leves),(30,total_onibus),(31,total_cam),(32,total_geral)]:
        c = ws.cell(row=tr, column=ci, value=val)
        c.font = Font(bold=True, color="1B5E20" if ci < 32 else "0D47A1", size=11)
        c.fill = hfill("E8F5E9"); c.alignment = ctr; c.border = brd

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(NCOLS)}{len(rows)+4}"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    nome = f"VDM_Trechos_{dt.now(BRASILIA).strftime('%d%m%Y')}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nome}"})



@app.get("/api/financeiro")
def get_financeiro():
    """Retorna dados financeiros para o painel da diretoria"""
    if not FINANCEIRO_PATH.exists():
        raise HTTPException(404, "Arquivo financeiro nao encontrado")

    df_all = pd.read_excel(FINANCEIRO_PATH, sheet_name="Controle de Gastos",
                           header=None, dtype=str)

    tipos   = ["COMBUSTIVEL","ALIMENTACAO VIAGEM","AGUA","OUTROS","HOSPEDAGEM","SERVICOS DE TERCEIROS"]
    viagens = []

    # Detecta viagens dinamicamente
    header_row = df_all.iloc[1]
    cols_viagem = []
    col_idx = 2
    while col_idx < len(df_all.columns):
        h = header_row.iloc[col_idx] if col_idx < len(header_row) else None
        if pd.notna(h) and str(h).strip() not in ('nan',''):
            viagens.append(str(h).strip())
            cols_viagem.append(col_idx)
        col_idx += 2
        if len(cols_viagem) > 0 and col_idx >= len(df_all.columns) - 2:
            break

    # Lê itens de gasto (linhas 3 a 8)
    itens = []
    for row_i in range(2, 8):
        row = df_all.iloc[row_i]
        tipo = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else tipos[row_i-2]
        vals = []
        for ci in cols_viagem:
            v = row.iloc[ci] if ci < len(row) else None
            try:
                fv = float(str(v).replace(",",".").strip())
                vals.append(0.0 if (fv != fv) else fv)
            except: vals.append(0.0)
        itens.append({"tipo": tipo, "valores": vals, "total": sum(vals)})

    # QTD pontos (linha 10)
    row_qtd = df_all.iloc[9]
    qtd_pontos = []
    for ci in cols_viagem:
        v = row_qtd.iloc[ci] if ci < len(row_qtd) else None
        try:
            fv = float(str(v).replace(",",".").strip())
            qtd_pontos.append(0 if (fv != fv) else int(fv))
        except: qtd_pontos.append(0)

    # Calcula totais por viagem
    n_viagens = len(viagens)
    totais_viagem_calc = [0.0] * n_viagens
    for item in itens:
        for j, v in enumerate(item["valores"]):
            if j < n_viagens:
                totais_viagem_calc[j] += v

    # Custo por ponto por viagem
    custo_por_ponto_viagem = []
    for j in range(n_viagens):
        qtd = qtd_pontos[j] if j < len(qtd_pontos) else 0
        cpp = round(totais_viagem_calc[j] / qtd, 2) if qtd > 0 else 0.0
        custo_por_ponto_viagem.append(cpp)

    total_geral = round(sum(totais_viagem_calc), 2)

    # Pontos concluídos por mês (baseado no período fim da aba DADOS COLETADOS)
    meses_pt = {1:"Janeiro",2:"Fevereiro",3:"Mar\u00e7o",4:"Abril",5:"Maio",6:"Junho",
                7:"Julho",8:"Agosto",9:"Setembro",10:"Outubro",11:"Novembro",12:"Dezembro"}
    pontos_por_mes = []
    try:
        from collections import defaultdict
        df_dc = pd.read_excel(EXCEL_PATH, sheet_name="DADOS COLETADOS", header=None)
        cont_mes = defaultdict(int)
        for i in range(4, len(df_dc)):
            row = df_dc.iloc[i]
            for col_d in [6, 5]:
                v = row.iloc[col_d]
                if pd.notna(v):
                    try:
                        ts = pd.Timestamp(v)
                        if ts.year > 1950:
                            cont_mes[(ts.year, ts.month)] += 1
                            break
                    except: pass
        for (ano, mes), qtd in sorted(cont_mes.items()):
            pontos_por_mes.append({
                "mes":   meses_pt[mes],
                "ano":   ano,
                "label": f"{meses_pt[mes]}/{ano}",
                "qtd":   qtd
            })
    except Exception as e:
        print(f"Erro pontos por mes: {e}")

    # Total de pontos = somatório dos meses (com data fim)
    total_pontos_mes = sum(m["qtd"] for m in pontos_por_mes) if pontos_por_mes else 0
    total_pontos = total_pontos_mes if total_pontos_mes > 0 else sum(qtd_pontos)
    custo_por_ponto = round(total_geral / total_pontos, 2) if total_pontos > 0 else 0.0

    return JSONResponse({
        "viagens":                viagens,
        "itens":                  itens,
        "totais_viagem":          [round(v, 2) for v in totais_viagem_calc],
        "qtd_pontos":             qtd_pontos,
        "custo_por_ponto_viagem": custo_por_ponto_viagem,
        "total_geral":            total_geral,
        "total_pontos":           total_pontos,
        "custo_por_ponto":        custo_por_ponto,
        "pontos_por_mes":         pontos_por_mes,
    })


@app.get("/api/ponto/{pid}")
def get_ponto(pid: str):
    pontos = ler_excel()
    for p in pontos:
        if p["id"] == pid: return JSONResponse(p)
    raise HTTPException(404, f"Ponto '{pid}' não encontrado")

def encontrar_xlsx_drive(pid: str) -> Optional[str]:
    """Encontra o arquivo .xlsx na pasta do ponto no Drive e retorna URL de download."""
    now = time.time()
    # Reutiliza cache de subpastas
    if "root_items" not in _cache_subpastas or (now - _cache_timestamp.get("root", 0)) > CACHE_TTL:
        items = drive_list(DRIVE_ROOT_FOLDER)
        _cache_subpastas["root_items"] = {
            i["name"]: i["id"] for i in items
            if i["mimeType"] == "application/vnd.google-apps.folder"
        }
        _cache_timestamp["root"] = now
    pasta_ponto_id = _cache_subpastas["root_items"].get(pid)
    if not pasta_ponto_id:
        return None
    # Lista arquivos direto na pasta do ponto (não subpasta)
    items = drive_list(pasta_ponto_id)
    for item in items:
        if (item.get("mimeType","") in
            {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
             "application/vnd.ms-excel"}
            or item.get("name","").endswith(".xlsx")):
            return f"https://drive.google.com/uc?export=download&id={item['id']}"
    return None

@app.get("/api/fotos/{pid}")
def get_fotos(pid: str):
    """Retorna fotos e link do xlsx de um ponto sob demanda"""
    fotos = listar_fotos_drive(pid)
    xlsx  = encontrar_xlsx_drive(pid)
    return JSONResponse({"pid": pid, "fotos": fotos, "total": len(fotos), "xlsx": xlsx})

@app.post("/api/upload-excel")
async def upload_excel(file: UploadFile = File(...)):
    if not file.filename.endswith((".xlsx",".xls")):
        raise HTTPException(400, "Apenas .xlsx aceito")
    EXCEL_PATH.write_bytes(await file.read())
    # Limpa cache de fotos
    _cache_fotos.clear()
    _cache_subpastas.clear()
    _cache_timestamp.clear()
    return {"ok": True, "atualizado_em": datetime.now(BRASILIA).strftime("%d/%m/%Y %H:%M:%S")}

@app.get("/api/limpar-cache")
def limpar_cache():
    """Força releitura das fotos do Drive"""
    _cache_fotos.clear()
    _cache_subpastas.clear()
    _cache_timestamp.clear()
    return {"ok": True, "msg": "Cache limpo"}

@app.get("/api/download/zip-xlsx")
def download_zip_xlsx():
    """Baixa ZIP com todos os xlsx dos pontos concluidos — download paralelo"""
    from fastapi.responses import StreamingResponse
    import unicodedata

    pontos = ler_excel()
    concluidos = []
    for p in pontos:
        k = unicodedata.normalize("NFD", p.get("status","").upper())
        k = "".join(c for c in k if unicodedata.category(c) != "Mn")
        if "CONCLU" in k:
            concluidos.append(p)

    if not concluidos:
        raise HTTPException(404, "Nenhum ponto concluido encontrado")

    # Garante cache do root
    now = time.time()
    if "root_items" not in _cache_subpastas or (now - _cache_timestamp.get("root", 0)) > CACHE_TTL:
        items = drive_list(DRIVE_ROOT_FOLDER)
        _cache_subpastas["root_items"] = {
            i["name"]: i["id"] for i in items
            if i["mimeType"] == "application/vnd.google-apps.folder"
        }
        _cache_timestamp["root"] = now

    def baixar_ponto(p):
        """Retorna (nome_arquivo, bytes) ou None"""
        pid = p["id"]
        pasta_id = _cache_subpastas["root_items"].get(pid)
        if not pasta_id:
            return None
        arquivos = drive_list(pasta_id)
        for arq in arquivos:
            mime = arq.get("mimeType", "")
            nome = arq.get("name", "")
            if (mime in {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         "application/vnd.ms-excel"}
                    or nome.lower().endswith(".xlsx")):
                try:
                    url = f"{DRIVE_BASE_URL}/files/{arq['id']}?alt=media&key={DRIVE_API_KEY}"
                    req = urllib.request.urlopen(url, timeout=30)
                    dados = req.read()
                    return (f"{pid}/{nome}", dados)
                except Exception as e:
                    print(f"Erro {pid}: {e}")
        return None

    # Download paralelo — até 8 pontos simultaneamente
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        with ThreadPoolExecutor(max_workers=8) as executor:
            futures = {executor.submit(baixar_ponto, p): p for p in concluidos}
            for future in as_completed(futures):
                result = future.result()
                if result:
                    nome_arq, dados = result
                    zf.writestr(nome_arq, dados)

    zip_buffer.seek(0)
    from datetime import datetime as dt
    nome_zip = f"Contagens_Concluidas_{dt.now(BRASILIA).strftime('%d%m%Y')}.zip"
    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename={nome_zip}"}
    )


@app.get("/api/download/resumo-xlsx")
def download_resumo():
    """Gera planilha resumo dos pontos concluidos"""
    from fastapi.responses import StreamingResponse
    from openpyxl.utils import get_column_letter
    from datetime import datetime as dt
    import unicodedata

    pontos = ler_excel()
    concluidos = []
    for p in pontos:
        k = unicodedata.normalize("NFD", p.get("status","").upper())
        k = "".join(c for c in k if unicodedata.category(c) != "Mn")
        if "CONCLU" in k:
            concluidos.append(p)

    if not concluidos:
        raise HTTPException(404, "Nenhum ponto concluido encontrado")

    wb = Workbook()
    ws = wb.active
    ws.title = "Pontos Concluidos"

    header_fill = PatternFill("solid", fgColor="1565C0")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left        = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin        = Side(style="thin", color="CCCCCC")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    green_fill  = PatternFill("solid", fgColor="E8F5E9")
    green_font  = Font(bold=True, color="1B5E20", size=11)

    NCOLS = 10

    ws.merge_cells(f"A1:{get_column_letter(NCOLS)}1")
    ws["A1"] = "RELATORIO DE PONTOS DE CONTAGEM CONCLUIDOS"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=13)
    ws["A1"].fill = PatternFill("solid", fgColor="0D47A1")
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f"A2:{get_column_letter(NCOLS)}2")
    ws["A2"] = f"COMOL Consultoria M.L.  |  SOP-CE  |  Gerado em: {dt.now(BRASILIA).strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(italic=True, color="555555", size=10)
    ws["A2"].fill = PatternFill("solid", fgColor="E3F2FD")
    ws["A2"].alignment = center
    ws.row_dimensions[2].height = 18

    headers = ["N", "ID Ponto", "Codigo Trecho", "Rodovia",
               "Descricao Inicio", "Descricao Fim",
               "Latitude", "Longitude",
               "Periodo Inicio", "Periodo Fim"]
    ws.row_dimensions[3].height = 36
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    widths = [5, 10, 16, 10, 36, 36, 13, 13, 13, 13]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    alt_fill = PatternFill("solid", fgColor="F5F5F5")
    for idx, p in enumerate(concluidos, 1):
        row = idx + 3
        ws.row_dimensions[row].height = 22
        fill = alt_fill if idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        valores = [
            idx,
            p.get("id", ""),
            p.get("trecho", ""),
            p.get("rodovia", ""),
            p.get("inicio", ""),
            p.get("fim", ""),
            p.get("lat", ""),
            p.get("lng", ""),
            p.get("periodo_inicio") or "-",
            p.get("periodo_fim")    or "-",
        ]
        for col, val in enumerate(valores, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill
            cell.border = border
            cell.alignment = center if col in [1,2,3,4,7,8,9,10] else left
            cell.font = Font(size=10)

    tot_row = len(concluidos) + 4
    ws.merge_cells(f"A{tot_row}:{get_column_letter(NCOLS)}{tot_row}")
    ws[f"A{tot_row}"] = f"TOTAL DE PONTOS CONCLUIDOS: {len(concluidos)}"
    ws[f"A{tot_row}"].font = green_font
    ws[f"A{tot_row}"].fill = green_fill
    ws[f"A{tot_row}"].alignment = center

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    nome = f"Resumo_Contagens_{dt.now(BRASILIA).strftime('%d%m%Y')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nome}"}
    )


@app.get("/api/financeiro")
def get_financeiro():
    """Retorna dados financeiros para o painel da diretoria"""
    if not FINANCEIRO_PATH.exists():
        raise HTTPException(404, "Arquivo financeiro nao encontrado")

    df_all = pd.read_excel(FINANCEIRO_PATH, sheet_name="Controle de Gastos",
                           header=None, dtype=str)

    tipos   = ["COMBUSTIVEL","ALIMENTACAO VIAGEM","AGUA","OUTROS","HOSPEDAGEM","SERVICOS DE TERCEIROS"]
    viagens = []

    # Detecta viagens dinamicamente (colunas de valores: 2, 4, 6, 8, 10...)
    header_row = df_all.iloc[1]
    cols_viagem = []
    col_idx = 2
    while col_idx < len(df_all.columns):
        h = header_row.iloc[col_idx] if col_idx < len(header_row) else None
        if pd.notna(h) and str(h).strip() not in ('nan',''):
            nome = str(h).strip()
            viagens.append(nome)
            cols_viagem.append(col_idx)
        col_idx += 2
        # Para de buscar se encontrou coluna TOTAL (ultima)
        if len(cols_viagem) > 0 and col_idx >= len(df_all.columns) - 2:
            break

    # Lê dados (linhas 3 a 8, índices 2 a 7)
    itens = []
    for row_i in range(2, 8):
        row = df_all.iloc[row_i]
        tipo = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else tipos[row_i-2]
        vals = []
        for col_idx in cols_viagem:
            v = row.iloc[col_idx] if col_idx < len(row) else None
            try:
                fv = float(str(v).replace(",",".").strip())
                vals.append(0.0 if (fv != fv) else fv)
            except: vals.append(0.0)
        itens.append({"tipo": tipo, "valores": vals, "total": sum(vals)})

    # Totais por viagem (linha 9)
    row_tot = df_all.iloc[8]
    totais_viagem = []
    for col_idx in cols_viagem:
        v = row_tot.iloc[col_idx] if col_idx < len(row_tot) else None
        try:
            fv = float(str(v).replace(",",".").strip())
            totais_viagem.append(0.0 if (fv != fv) else fv)
        except: totais_viagem.append(0.0)

    # QTD pontos (linha 10)
    row_qtd = df_all.iloc[9]
    qtd_pontos = []
    for col_idx in cols_viagem:
        v = row_qtd.iloc[col_idx] if col_idx < len(row_qtd) else None
        try:
            fv = float(str(v).replace(",",".").strip())
            qtd_pontos.append(0 if (fv != fv) else int(fv))
        except: qtd_pontos.append(0)



    # Calcula totais por viagem somando os itens (ignora fórmulas do Excel)
    n_viagens = len(viagens)
    totais_viagem_calc = [0.0] * n_viagens
    for item in itens:
        for j, v in enumerate(item["valores"]):
            if j < n_viagens:
                totais_viagem_calc[j] += v

    # Custo por ponto por viagem
    custo_por_ponto_viagem = []
    for j in range(3):
        cpp = round(totais_viagem_calc[j] / qtd_pontos[j], 2) if qtd_pontos[j] > 0 else 0.0
        custo_por_ponto_viagem.append(cpp)

    # Totais gerais
    total_geral  = round(sum(totais_viagem_calc), 2)
    # Total de pontos = somatório dos meses (apenas pontos com data fim preenchida)
    total_pontos_mes = sum(m["qtd"] for m in pontos_por_mes) if pontos_por_mes else 0
    total_pontos = total_pontos_mes if total_pontos_mes > 0 else sum(qtd_pontos)
    custo_por_ponto = round(total_geral / total_pontos, 2) if total_pontos > 0 else 0.0


    return JSONResponse({
        "total": len(pontos),
        "com_contagem": sum(1 for p in pontos if p["total"] > 0),
        "atualizado_em": excel_mtime,
        "pontos": pontos,
    })

@app.get("/api/ponto/{pid}")
def get_ponto(pid: str):
    pontos = ler_excel()
    for p in pontos:
        if p["id"] == pid: return JSONResponse(p)
    raise HTTPException(404, f"Ponto '{pid}' não encontrado")

def encontrar_xlsx_drive(pid: str) -> Optional[str]:
    """Encontra o arquivo .xlsx na pasta do ponto no Drive e retorna URL de download."""
    now = time.time()
    # Reutiliza cache de subpastas
    if "root_items" not in _cache_subpastas or (now - _cache_timestamp.get("root", 0)) > CACHE_TTL:
        items = drive_list(DRIVE_ROOT_FOLDER)
        _cache_subpastas["root_items"] = {
            i["name"]: i["id"] for i in items
            if i["mimeType"] == "application/vnd.google-apps.folder"
        }
        _cache_timestamp["root"] = now
    pasta_ponto_id = _cache_subpastas["root_items"].get(pid)
    if not pasta_ponto_id:
        return None
    # Lista arquivos direto na pasta do ponto (não subpasta)
    items = drive_list(pasta_ponto_id)
    for item in items:
        if (item.get("mimeType","") in
            {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
             "application/vnd.ms-excel"}
            or item.get("name","").endswith(".xlsx")):
            return f"https://drive.google.com/uc?export=download&id={item['id']}"
    return None

@app.get("/api/fotos/{pid}")
def get_fotos(pid: str):
    """Retorna fotos e link do xlsx de um ponto sob demanda"""
    fotos = listar_fotos_drive(pid)
    xlsx  = encontrar_xlsx_drive(pid)
    return JSONResponse({"pid": pid, "fotos": fotos, "total": len(fotos), "xlsx": xlsx})

@app.post("/api/upload-excel")
async def upload_excel(file: UploadFile = File(...)):
    if not file.filename.endswith((".xlsx",".xls")):
        raise HTTPException(400, "Apenas .xlsx aceito")
    EXCEL_PATH.write_bytes(await file.read())
    # Limpa cache de fotos
    _cache_fotos.clear()
    _cache_subpastas.clear()
    _cache_timestamp.clear()
    return {"ok": True, "atualizado_em": datetime.now(BRASILIA).strftime("%d/%m/%Y %H:%M:%S")}

@app.get("/api/limpar-cache")
def limpar_cache():
    """Força releitura das fotos do Drive"""
    _cache_fotos.clear()
    _cache_subpastas.clear()
    _cache_timestamp.clear()
    return {"ok": True, "msg": "Cache limpo"}

@app.get("/api/download/zip-xlsx")
def download_zip_xlsx():
    """Baixa ZIP com todos os xlsx dos pontos concluidos — download paralelo"""
    from fastapi.responses import StreamingResponse
    import unicodedata

    pontos = ler_excel()
    concluidos = []
    for p in pontos:
        k = unicodedata.normalize("NFD", p.get("status","").upper())
        k = "".join(c for c in k if unicodedata.category(c) != "Mn")
        if "CONCLU" in k:
            concluidos.append(p)

    if not concluidos:
        raise HTTPException(404, "Nenhum ponto concluido encontrado")

    # Garante cache do root
    now = time.time()
    if "root_items" not in _cache_subpastas or (now - _cache_timestamp.get("root", 0)) > CACHE_TTL:
        items = drive_list(DRIVE_ROOT_FOLDER)
        _cache_subpastas["root_items"] = {
            i["name"]: i["id"] for i in items
            if i["mimeType"] == "application/vnd.google-apps.folder"
        }
        _cache_timestamp["root"] = now

    def baixar_ponto(p):
        """Retorna (nome_arquivo, bytes) ou None"""
        pid = p["id"]
        pasta_id = _cache_subpastas["root_items"].get(pid)
        if not pasta_id:
            return None
        arquivos = drive_list(pasta_id)
        for arq in arquivos:
            mime = arq.get("mimeType", "")
            nome = arq.get("name", "")
            if (mime in {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         "application/vnd.ms-excel"}
                    or nome.lower().endswith(".xlsx")):
                try:
                    url = f"{DRIVE_BASE_URL}/files/{arq['id']}?alt=media&key={DRIVE_API_KEY}"
                    req = urllib.request.urlopen(url, timeout=30)
                    dados = req.read()
                    return (f"{pid}/{nome}", dados)
                except Exception as e:
                    print(f"Erro {pid}: {e}")
        return None

    # Download paralelo — até 8 pontos simultaneamente
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        with ThreadPoolExecutor(max_workers=8) as executor:
            futures = {executor.submit(baixar_ponto, p): p for p in concluidos}
            for future in as_completed(futures):
                result = future.result()
                if result:
                    nome_arq, dados = result
                    zf.writestr(nome_arq, dados)

    zip_buffer.seek(0)
    from datetime import datetime as dt
    nome_zip = f"Contagens_Concluidas_{dt.now(BRASILIA).strftime('%d%m%Y')}.zip"
    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename={nome_zip}"}
    )


@app.get("/api/download/resumo-xlsx")
def download_resumo():
    """Gera planilha resumo dos pontos concluidos"""
    from fastapi.responses import StreamingResponse
    from openpyxl.utils import get_column_letter
    from datetime import datetime as dt
    import unicodedata

    pontos = ler_excel()
    concluidos = []
    for p in pontos:
        k = unicodedata.normalize("NFD", p.get("status","").upper())
        k = "".join(c for c in k if unicodedata.category(c) != "Mn")
        if "CONCLU" in k:
            concluidos.append(p)

    if not concluidos:
        raise HTTPException(404, "Nenhum ponto concluido encontrado")

    wb = Workbook()
    ws = wb.active
    ws.title = "Pontos Concluidos"

    header_fill = PatternFill("solid", fgColor="1565C0")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left        = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin        = Side(style="thin", color="CCCCCC")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    green_fill  = PatternFill("solid", fgColor="E8F5E9")
    green_font  = Font(bold=True, color="1B5E20", size=11)

    NCOLS = 10

    ws.merge_cells(f"A1:{get_column_letter(NCOLS)}1")
    ws["A1"] = "RELATORIO DE PONTOS DE CONTAGEM CONCLUIDOS"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=13)
    ws["A1"].fill = PatternFill("solid", fgColor="0D47A1")
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f"A2:{get_column_letter(NCOLS)}2")
    ws["A2"] = f"COMOL Consultoria M.L.  |  SOP-CE  |  Gerado em: {dt.now(BRASILIA).strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(italic=True, color="555555", size=10)
    ws["A2"].fill = PatternFill("solid", fgColor="E3F2FD")
    ws["A2"].alignment = center
    ws.row_dimensions[2].height = 18

    headers = ["N", "ID Ponto", "Codigo Trecho", "Rodovia",
               "Descricao Inicio", "Descricao Fim",
               "Latitude", "Longitude",
               "Periodo Inicio", "Periodo Fim"]
    ws.row_dimensions[3].height = 36
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    widths = [5, 10, 16, 10, 36, 36, 13, 13, 13, 13]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    alt_fill = PatternFill("solid", fgColor="F5F5F5")
    for idx, p in enumerate(concluidos, 1):
        row = idx + 3
        ws.row_dimensions[row].height = 22
        fill = alt_fill if idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        valores = [
            idx,
            p.get("id", ""),
            p.get("trecho", ""),
            p.get("rodovia", ""),
            p.get("inicio", ""),
            p.get("fim", ""),
            p.get("lat", ""),
            p.get("lng", ""),
            p.get("periodo_inicio") or "-",
            p.get("periodo_fim")    or "-",
        ]
        for col, val in enumerate(valores, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill
            cell.border = border
            cell.alignment = center if col in [1,2,3,4,7,8,9,10] else left
            cell.font = Font(size=10)

    tot_row = len(concluidos) + 4
    ws.merge_cells(f"A{tot_row}:{get_column_letter(NCOLS)}{tot_row}")
    ws[f"A{tot_row}"] = f"TOTAL DE PONTOS CONCLUIDOS: {len(concluidos)}"
    ws[f"A{tot_row}"].font = green_font
    ws[f"A{tot_row}"].fill = green_fill
    ws[f"A{tot_row}"].alignment = center

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    nome = f"Resumo_Contagens_{dt.now(BRASILIA).strftime('%d%m%Y')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nome}"}
    )


@app.get("/api/financeiro")
def get_financeiro():
    """Retorna dados financeiros para o painel da diretoria"""
    if not FINANCEIRO_PATH.exists():
        raise HTTPException(404, "Arquivo financeiro nao encontrado")

    df_all = pd.read_excel(FINANCEIRO_PATH, sheet_name="Controle de Gastos",
                           header=None, dtype=str)

    tipos   = ["COMBUSTIVEL","ALIMENTACAO VIAGEM","AGUA","OUTROS","HOSPEDAGEM","SERVICOS DE TERCEIROS"]
    viagens = []

    # Detecta viagens dinamicamente (colunas de valores: 2, 4, 6, 8, 10...)
    header_row = df_all.iloc[1]
    cols_viagem = []
    col_idx = 2
    while col_idx < len(df_all.columns):
        h = header_row.iloc[col_idx] if col_idx < len(header_row) else None
        if pd.notna(h) and str(h).strip() not in ('nan',''):
            nome = str(h).strip()
            viagens.append(nome)
            cols_viagem.append(col_idx)
        col_idx += 2
        # Para de buscar se encontrou coluna TOTAL (ultima)
        if len(cols_viagem) > 0 and col_idx >= len(df_all.columns) - 2:
            break

    # Lê dados (linhas 3 a 8, índices 2 a 7)
    itens = []
    for row_i in range(2, 8):
        row = df_all.iloc[row_i]
        tipo = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else tipos[row_i-2]
        vals = []
        for col_idx in cols_viagem:
            v = row.iloc[col_idx] if col_idx < len(row) else None
            try:
                fv = float(str(v).replace(",",".").strip())
                vals.append(0.0 if (fv != fv) else fv)
            except: vals.append(0.0)
        itens.append({"tipo": tipo, "valores": vals, "total": sum(vals)})

    # Totais por viagem (linha 9)
    row_tot = df_all.iloc[8]
    totais_viagem = []
    for col_idx in cols_viagem:
        v = row_tot.iloc[col_idx] if col_idx < len(row_tot) else None
        try:
            fv = float(str(v).replace(",",".").strip())
            totais_viagem.append(0.0 if (fv != fv) else fv)
        except: totais_viagem.append(0.0)

    # QTD pontos (linha 10)
    row_qtd = df_all.iloc[9]
    qtd_pontos = []
    for col_idx in cols_viagem:
        v = row_qtd.iloc[col_idx] if col_idx < len(row_qtd) else None
        try:
            fv = float(str(v).replace(",",".").strip())
            qtd_pontos.append(0 if (fv != fv) else int(fv))
        except: qtd_pontos.append(0)

    # Calcula totais por viagem somando os itens (ignora fórmulas do Excel)
    n_viagens = len(viagens)
    totais_viagem_calc = [0.0] * n_viagens
    for item in itens:
        for j, v in enumerate(item["valores"]):
            if j < n_viagens:
                totais_viagem_calc[j] += v

    # Custo por ponto por viagem
    custo_por_ponto_viagem = []
    for j in range(3):
        cpp = round(totais_viagem_calc[j] / qtd_pontos[j], 2) if qtd_pontos[j] > 0 else 0.0
        custo_por_ponto_viagem.append(cpp)

    # Totais gerais
    total_geral  = round(sum(totais_viagem_calc), 2)
    # Total de pontos = somatório dos meses (apenas pontos com data fim preenchida)
    total_pontos_mes = sum(m["qtd"] for m in pontos_por_mes) if pontos_por_mes else 0
    total_pontos = total_pontos_mes if total_pontos_mes > 0 else sum(qtd_pontos)
    custo_por_ponto = round(total_geral / total_pontos, 2) if total_pontos > 0 else 0.0

    # Pontos concluídos por mês (baseado no período fim da aba DADOS COLETADOS)
    meses_pt = {1:"Janeiro",2:"Fevereiro",3:"Mar\u00e7o",4:"Abril",5:"Maio",6:"Junho",
                7:"Julho",8:"Agosto",9:"Setembro",10:"Outubro",11:"Novembro",12:"Dezembro"}
    pontos_por_mes = []
    try:
        from collections import defaultdict
        df_dc = pd.read_excel(EXCEL_PATH, sheet_name="DADOS COLETADOS", header=None)
        cont_mes = defaultdict(int)
        for i in range(4, len(df_dc)):
            row = df_dc.iloc[i]
            # Tenta período fim (col 6), se inválida usa início (col 5)
            data_encontrada = False
            for col_idx in [6, 5]:
                v = row.iloc[col_idx]
                if pd.notna(v):
                    try:
                        ts = pd.Timestamp(v)
                        if ts.year > 1950:
                            cont_mes[(ts.year, ts.month)] += 1
                            data_encontrada = True
                            break
                    except: pass
            _ = data_encontrada  # evita warning
        for (ano, mes), qtd in sorted(cont_mes.items()):
            pontos_por_mes.append({
                "mes":   meses_pt[mes],
                "ano":   ano,
                "label": f"{meses_pt[mes]}/{ano}",
                "qtd":   qtd
            })
    except Exception as e:
        print(f"Erro pontos por mes: {e}")

    return JSONResponse({
        "viagens":               viagens,
        "itens":                 itens,
        "totais_viagem":         [round(v, 2) for v in totais_viagem_calc],
        "qtd_pontos":            qtd_pontos,
        "custo_por_ponto_viagem": custo_por_ponto_viagem,
        "total_geral":           total_geral,
        "total_pontos":          total_pontos,
        "custo_por_ponto":       custo_por_ponto,
        "pontos_por_mes":        pontos_por_mes,
    })


@app.get("/api/status")
def status():
    return {
        "ok": True,
        "excel_existe": EXCEL_PATH.exists(),
        "excel_modificado": datetime.fromtimestamp(EXCEL_PATH.stat().st_mtime, tz=BRASILIA).strftime("%d/%m/%Y %H:%M:%S") if EXCEL_PATH.exists() else None,
        "drive_root": DRIVE_ROOT_FOLDER,
        "cache_pontos": len(_cache_fotos),
        "servidor_hora": datetime.now(BRASILIA).strftime("%d/%m/%Y %H:%M:%S"),
    }
